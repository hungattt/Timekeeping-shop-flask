import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from functools import wraps
from flask import Flask, flash, render_template, request, jsonify, redirect, url_for, session, send_file
import logging
import threading
# create logger with 'my_app'
logger = logging.getLogger('KT')
logger.setLevel(logging.DEBUG)

logging.basicConfig(filename="log.txt", level=logging.ERROR, encoding="utf-8")
# logging.basicConfig(handlers=[logging.FileHandler(filename="./log.txt",
#                                                  encoding='utf-8', mode='a+')],
#                     format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
#                     datefmt="%F %A %T",
#                     level=logging.INFO)


excel_lock = threading.Lock()
app = Flask(__name__)
app.secret_key = b'_3#y6L"F9Q3z\n\xec]/'

# setup
sheetname = "sum"
data_log_name = "input/data_all.xlsx"


# Read account file to check user login
workbook = load_workbook("input/data_all.xlsx")
dataSheet = workbook['sum']

# Get the current date and time
now = datetime.datetime.now()

# Read the Excel file ( data frame )
df = pd.read_excel('input/account.xlsx')
df.set_index('userId', inplace=True)

workbook = Workbook()
# Create a new sheet and make it visible
sheet = workbook.active
sheet.title = 'Sheet1'

# read Excel file account
excel_account = pd.read_excel('input/data_all.xlsx')
# read timelog excel
excel_timelog = pd.read_excel(data_log_name, sheet_name=sheetname)

# Set the Ho ten column as the index
excel_timelog.set_index('notduplicate', inplace=True)


@app.route('/', methods=['GET'])
def index():
    return render_template('welcome.html')


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' in session:
            return f(*args, **kwargs)
        else:
            return redirect(url_for('login'))
    return decorated_function


@app.route('/hello', methods=['GET'])
def hello():
    if 'user_login_info' in session:
        return f'Logged in as {session["user_login_info"]}'
    return 'You are not logged in'


@app.route('/getlog', methods=['GET'])
@login_required
def getlog():
    return send_file('log.txt', mimetype='text/plain')


@app.route('/getdata', methods=['GET'])
@login_required
def getdata():
    if 'user_login_info' in session:
        return send_file(data_log_name, as_attachment=True)
    return 'You are not logged in'


@app.route('/count', methods=['GET', 'POST'])
def count():
    if 'user_login_info' in session:
        if 'count_money' in session:
            count_money = session["count_money"]
        else:
            session["count_money"] = 0
        return render_template('index.html', count_money=count_money)
    return 'You are not logged in'


def get_money_final_and_print_log(input_data, excel_data, str_log, input_type, user_login_id):
    if input_data != "":
        month_input = input_data.replace(",", "")
        month_final = int(excel_data) + int(month_input)
        str_data = str_log + input_type + \
            "{:,.0f}".format(month_final) + \
            " (+" + "{:,.0f}".format(int(month_input)) + ") "
    else:
        month_input = 0
        month_final = excel_data
        str_data = str_log + input_type + "{:,.0f}".format(month_final)
    print(user_login_id + ": " + str_data)
    logger.info(user_login_id + ": " + str_data)
    return month_final


def change_money_by_userid(userId, month, week, thanks, celebrate):
    global excel_lock
    with excel_lock:
        workbook = load_workbook("input/data_all.xlsx")
        dataSheet = workbook['sum']
        msg_name = ""
        idRow = int(userId.split("_")[0]) + 1
        try:
            row_userid = dataSheet.cell(row=idRow, column=2).value
            row_name = dataSheet.cell(row=idRow, column=3).value
            row_khoa = dataSheet.cell(row=idRow, column=1).value
            row_month = 0 if dataSheet.cell(row=idRow, column=5).value is None else int(
                dataSheet.cell(row=idRow, column=5).value)
            row_week = 0 if dataSheet.cell(row=idRow, column=6).value is None else int(
                dataSheet.cell(row=idRow, column=6).value)
            row_thanks = 0 if dataSheet.cell(row=idRow, column=7).value is None else int(
                dataSheet.cell(row=idRow, column=7).value)
            row_celebrate = 0 if dataSheet.cell(row=idRow, column=8).value is None else int(
                dataSheet.cell(row=idRow, column=8).value)
            now = datetime.datetime.now()
            str_data = str(now) + ": " + str(row_khoa) + "_" + row_name + ': '

            str_login_append = session["user_login_info"]
            user_login_col = session["user_login_info"] if dataSheet.cell(row=idRow, column=9).value is None else str(
            dataSheet.cell(row=idRow, column=9).value) + " | " + session["user_login_info"]

            msg_name = row_name + ", Khoá " + str(row_khoa)
            print("Bắt đầu xử lý: " + str(row_userid) +
                  " | " + "Anh em = " + msg_name)
            # Version no
            row_ver_no = 1 if dataSheet.cell(row=idRow, column=10).value is None else int(dataSheet.cell(row=idRow, column=10).value) + 1
            dataSheet.cell(row=idRow, column=9).value = user_login_col
            dataSheet.cell(row=idRow, column=10).value = row_ver_no
            print("Version_no: " + str(row_ver_no))
            logger.info("Version_no: " + str(row_ver_no))

            # Month
            month_final = get_money_final_and_print_log(
                month, row_month, str_data, "Month: ", str_login_append)
            dataSheet.cell(row=idRow, column=5).value = month_final

            # week
            week_final = get_money_final_and_print_log(
                week, row_week, str_data, "Week: ", str_login_append)
            dataSheet.cell(row=idRow, column=6).value = week_final

            # thanks
            thanks_final = get_money_final_and_print_log(
                thanks, row_thanks, str_data, "Thanks: ", str_login_append)
            dataSheet.cell(row=idRow, column=7).value = thanks_final

            # Month
            celebrate_final = get_money_final_and_print_log(
                celebrate, row_celebrate, str_data, "Celebrate: ", str_login_append)
            dataSheet.cell(row=idRow, column=8).value = celebrate_final

            workbook.save("input/data_all.xlsx")
            workbook.close()

        except Exception:
            print("Lỗi khi xử ký data: " + userId)
            logger.error("Lỗi khi xử ký data: " + userId)

    return msg_name


@app.route('/form', methods=['GET', 'POST'])
@login_required
def form():
    if request.method == 'GET':
        # Lấy thuộc tính từ parameter URL
        key = request.args.get('key')
        return render_template('form.html', key=key)
    elif request.method == 'POST':
        # Get data from view
        month = request.form['month']
        week = request.form['week']
        thanks = request.form['thanks']
        celebrate = request.form['celebrate']
        userId = request.form['key']
        logger.info("Bắt đầu xử lý ID=" + userId)
        idName = change_money_by_userid(userId, month, week, thanks, celebrate)

        # Set form_submitted variable to True
        form_submitted = True
        if idName == "":
            msg = "Không tìm thấy anh em tương ứng trong danh sách"
        else:
            msg = idName + ": "
            if len(month) != 0:
                msg = msg + "Tháng: " + month
            if len(week) != 0:
                msg = msg + "| Tuần: " + week
            if len(thanks) != 0:
                msg = msg + "| Cảm tạ: " + thanks
            if len(celebrate) != 0:
                msg = msg + "| Kỷ niệm: " + celebrate

        # Chuyển hướng hoặc hiển thị trang thành công
        return render_template('form.html', form_submitted=form_submitted, msg=msg)
        # return redirect(url_for('search'))


@app.route('/search', methods=['GET', 'POST'])
@login_required
def search():
    if request.method == 'GET':
        # list account
        mylist = []

        # Sử dụng openpyxl
        for row in dataSheet.iter_rows(min_row=2, values_only=True):
            hoten = str(row[2])
            korea = str(row[3])
            khoa = str(row[0])
            stt = str(row[1])
            notduplicate = stt + "_" + hoten + "_" + khoa + "_" + korea
            mylist.append(notduplicate)

        return render_template('index.html', data=mylist)
    elif request.method == 'POST':
        keyword = request.form['keyword']
        # connect to database and query for data based on keyword
        # store the data in a variable, such as result
        return render_template('l', result=keyword)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        print(email + "_" + password)

        try:
            # attempt to get the row with the given email
            user_obj = df.loc[email]
            # print the result
            print(user_obj)
            if user_obj is not None:
                row_userid = email
                row_password = user_obj.password

                print("data row: " + row_userid + "_" + row_password)
                if password == row_password:
                    # Xóa thông báo lỗi sau khi hiển thị
                    session.pop('error_message', None)
                    # set session login
                    session['logged_in'] = True
                    session['user_login_info'] = email
                    # Đăng nhập thành công, chuyển hướng hoặc trả về một trang HTML
                    return redirect(url_for('search'))

        except KeyError:
            # handle the case when the email is not in the index
            print(f"No user with email {email} found.")
            logger.error(f"No user with email {email} found.")
        except Exception as e:
            # handle any other unexpected errors
            print(f"Lỗi không xác định, liên hệ admin: {e}")
            logger.error(f"Lỗi không xác định, liên hệ admin: {e}")

        # Lưu thông báo lỗi vào session
        session['error_message'] = 'Tên người dùng hoặc mật khẩu không chính xác'
        return render_template('login.html')

    # Xóa thông báo lỗi sau khi hiển thị
    session.pop('error_message', None)

    # Trả về trang HTML đăng nhập
    return render_template('login.html')


@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    if request.method == 'POST' or request.method == 'GET':
        session.pop('logged_in', None)
        session.pop('error_message', None)
        session.pop('user_login_info', None)
        return render_template('login.html')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3693, debug=False)
